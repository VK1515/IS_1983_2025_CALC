import streamlit as st
import pandas as pd
import math
import matplotlib.pyplot as plt
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, Image
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# ==================================================
# PAGE CONFIG
# ==================================================
st.set_page_config(
    page_title="IS 1893:2025 Seismic Force Calculator",
    layout="wide"
)

st.title("IS 1893:2025 – Seismic Base Shear Calculator")
st.caption(
    "Equivalent Static Method | Multi-Zone Capability\n\n"
    
)

# ==================================================
# Z TABLE (IS 1893:2025)
# ==================================================
Z_TABLE = {
    "VI": {75:0.300,175:0.375,275:0.450,475:0.500,975:0.600,1275:0.625,2475:0.750,4975:0.940,9975:1.125},
    "V":  {75:0.200,175:0.250,275:0.300,475:0.333,975:0.400,1275:0.4167,2475:0.500,4975:0.625,9975:0.750},
    "IV": {75:0.140,175:0.175,275:0.210,475:0.233,975:0.280,1275:0.2917,2475:0.350,4975:0.440,9975:0.525},
    "III":{75:0.0625,175:0.085,275:0.100,475:0.125,975:0.167,1275:0.1875,2475:0.250,4975:0.333,9975:0.450},
    "II": {75:0.0375,175:0.050,275:0.060,475:0.075,975:0.100,1275:0.1125,2475:0.150,4975:0.200,9975:0.270}
}

# ==================================================
# SESSION STATE
# ==================================================
if "base_shear" not in st.session_state:
    st.session_state.base_shear = {}
if "multi_zone_df" not in st.session_state:
    st.session_state.multi_zone_df = None

# ==================================================
# FUNCTIONS
# ==================================================
def A_NH(T, site):
    if site == "A/B":
        return 2.5 if T <= 0.4 else (1/T if T <= 6 else 6/T**2)
    if site == "C":
        return 2.5 if T <= 0.6 else (1.5/T if T <= 6 else 9/T**2)
    return 2.5 if T <= 0.8 else (2/T if T <= 6 else 12/T**2)

def delta_v(T, site):
    if T > 0.10:
        return 0.67
    return {"A/B":0.80, "C":0.82, "D":0.85}[site]

def gamma_v(T, site):
    return {"A/B":1/T, "C":1.5/T, "D":2.0/T}[site]

# ==================================================
# TABS
# ==================================================
tab1,  tab3 = st.tabs([
    "① Base Shear (X & Y)",
    
    "③ Multi-Zone Study & Export"
])

# ==================================================
# TAB 1 – BASE SHEAR
# ==================================================
with tab1:
    st.subheader("Direction-wise Base Shear Calculation")

    zone = st.selectbox("Earthquake Zone", ["II","III","IV","V","VI"])
    TR = st.selectbox("Return Period TR (years)", [75,175,275,475,975,1275,2475,4975,9975])
    Z = Z_TABLE[zone][TR]

    I = st.number_input("Importance Factor (I)", value=1.0)
    R = st.number_input("Response Reduction Factor (R)", value=5.0)
    site = st.selectbox("Site Class", ["A/B","C","D"])
    W = st.number_input("Total Seismic Weight W (kN)", value=10000.0)

    H = st.number_input("Total Height H (m)", value=15.0)
    dx = st.number_input("Plan Dimension dx (m)", value=10.0)
    dy = st.number_input("Plan Dimension dy (m)", value=15.0)
    TV = st.number_input("Vertical Period TV (s)", value=0.4)

    if st.button("Compute Base Shear"):
        Tx = 0.09 * H / math.sqrt(dx)
        Ty = 0.09 * H / math.sqrt(dy)

        Vx = (Z * I * A_NH(Tx, site) / R) * W
        Vy = (Z * I * A_NH(Ty, site) / R) * W
        Vv = Z * I * delta_v(TV, site) * gamma_v(TV, site) * W

        st.session_state.base_shear = {
            "Zone":zone,"TR":TR,"Z":Z,
            "I":I,"R":R,"Site":site,
            "W":W,"Tx":Tx,"Ty":Ty,
            "Vx":Vx,"Vy":Vy,"Vv":Vv
        }

        st.success("Base shear computed and locked.")

        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: st.metric("Tx (s)", f"{Tx:.3f}")
        with c2: st.metric("Ty (s)", f"{Ty:.3f}")
        with c3: st.metric("Vx (kN)", f"{Vx:.2f}")
        with c4: st.metric("Vy (kN)", f"{Vy:.2f}")
        with c5: st.metric("Vv (kN)", f"{Vv:.2f}")

# ==================================================






# ==================================================
# TAB 3 – MULTI-ZONE STUDY + EXPORT
# ==================================================
with tab3:
    st.subheader("Multi-Zone Base Shear Comparison")

    zones = st.multiselect("Select Zones", ["II","III","IV","V","VI"], default=["II","III","IV","V"])
    TR = st.selectbox("Return Period (years)", [75,175,275,475,975,1275,2475,4975,9975], key="mz_tr")
    I = st.number_input("Importance Factor", value=1.0, key="mz_I")
    R = st.number_input("Response Reduction Factor", value=5.0, key="mz_R")
    site = st.selectbox("Site Class", ["A/B","C","D"], key="mz_site")
    W = st.number_input("Total Weight W (kN)", value=10000.0, key="mz_W")
    H = st.number_input("Height H (m)", value=15.0, key="mz_H")
    dx = st.number_input("dx (m)", value=10.0, key="mz_dx")
    dy = st.number_input("dy (m)", value=15.0, key="mz_dy")

    if st.button("Compute Multi-Zone Base Shear"):
        Tx = 0.09 * H / math.sqrt(dx)
        Ty = 0.09 * H / math.sqrt(dy)

        data=[]
        for z in zones:
            Z = Z_TABLE[z][TR]
            data.append([z,Z,(Z*I*A_NH(Tx,site)/R)*W,(Z*I*A_NH(Ty,site)/R)*W])

        dfz = pd.DataFrame(data, columns=["Zone","Z","Vx (kN)","Vy (kN)"])
        st.session_state.multi_zone_df = dfz

        st.dataframe(dfz.round(3), use_container_width=True)

        fig, ax = plt.subplots()
        ax.plot(dfz["Zone"], dfz["Vx (kN)"], marker="o", label="X direction")
        ax.plot(dfz["Zone"], dfz["Vy (kN)"], marker="s", label="Y direction")
        ax.set_xlabel("Seismic Zone")
        ax.set_ylabel("Base Shear (kN)")
        ax.set_title("Base Shear vs Seismic Zone")
        ax.grid(True)
        ax.legend()
        st.pyplot(fig)

        fig.savefig("base_shear_zone.png", dpi=300)

    # ---------------- EXPORT ----------------
    if st.session_state.multi_zone_df is not None:
        dfz = st.session_state.multi_zone_df

        # Excel
        excel_file="IS1893_2025_BaseShear_MultiZone.xlsx"
        dfz.to_excel(excel_file,index=False)
        wb=load_workbook(excel_file)
        ws=wb.active

        chart=LineChart()
        chart.title="Base Shear vs Zone"
        chart.y_axis.title="Base Shear (kN)"
        chart.x_axis.title="Zone"
        data=Reference(ws,min_col=3,min_row=1,max_col=4,max_row=ws.max_row)
        cats=Reference(ws,min_col=1,min_row=2,max_row=ws.max_row)
        chart.add_data(data,titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart,"G2")
        wb.save(excel_file)

        st.download_button("Download Excel (with graph)", open(excel_file,"rb"), file_name=excel_file)

        # PDF
        pdf_file="IS1893_2025_BaseShear_MultiZone.pdf"
        doc=SimpleDocTemplate(pdf_file)
        styles=getSampleStyleSheet()
        content=[
            Paragraph("IS 1893:2025 – Multi-Zone Base Shear Study", styles["Title"]),
            Paragraph("For educational use only<br/>Created by: Vrushali Kamalakar", styles["Normal"]),
            Table([dfz.columns.tolist()]+dfz.round(3).values.tolist()),
            Image("base_shear_zone.png", width=400, height=250)
        ]
        doc.build(content)

        st.download_button("Download PDF (with graph)", open(pdf_file,"rb"), file_name=pdf_file)

# ==================================================
# FOOTER
# ==================================================
st.markdown("---")
st.info(
    "📘 **For Educational Use Only**\n\n"
    "Independent verification is mandatory before professional or statutory use.\n\n"
    "**Created by: Vrushali Kamalakar**"
)



